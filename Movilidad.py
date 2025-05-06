# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox, Toplevel, Listbox, Scrollbar, Button, Label, Frame, Entry, StringVar
import sys
import os
import unicodedata
import re
from thefuzz import process
import math
import traceback
from typing import List, Dict, Tuple, Optional, Any
import random # Importado para aleatoriedad


# ==============================================================================
# --- Constantes de Configuración ---
# ==============================================================================

# --- Nombres de Hojas ---
SOURCE_SHEET_NAME = 'WEB2.0'
TARGET_SHEET_NAME = 'Plantilla de Cargue 2'
LOOKUP_SUCURSAL_SHEET_NAME = 'Clientes y sucursales'
LOOKUP_CSAP1_SHEET_NAME = 'CSAP2'
LOOKUP_CSAP2_SHEET_NAME = 'CodigoSAP'

# --- Nombres de Columnas (Fuente: SOURCE_SHEET_NAME) ---
SRC_COL_CLIENTE = 'Cliente'
SRC_COL_FECHA = 'Fecha CC'
SRC_COL_PLACA = 'Placa'
SRC_COL_PESO = 'Peso Recogido'
SRC_COL_RESIDUO = 'Residuo'

# --- Nombres de Columnas (Destino: TARGET_SHEET_NAME) ---
TGT_COL_CLIENTE = 'CLIENTE'
TGT_COL_FECHA = 'FECHA DE SERVICIO'
TGT_COL_PLACA = 'PLACA DEL VEHÍCULO'
TGT_COL_PESO = 'CANTIDAD (PESO)'
TGT_COL_SUCURSAL = 'SUCURSAL'
TGT_COL_SAP = 'CÓDIGO SAP DEL MATERIAL'
TGT_COL_DEUDOR_SUC = 'DEUDOR DE SUCURSAL'
TGT_COL_CEDULA_CONDUCTOR = 'CÉDULA DE CONDUCTOR'
TGT_COL_CEDULA_AUXILIAR = 'CÉDULA DE AUXILIAR'       # <<< NUEVA COLUMNA
TGT_COL_NOMBRE_ENTREGA = 'NOMBRE DE QUIEN ENTREGA' # <<< NUEVA COLUMNA
TGT_COL_CARGO_ENTREGA = 'CARGO DE QUIEN ENTREGA'   # <<< NUEVA COLUMNA
TGT_COL_UNO = '1'                                  # <<< Nombre de columna duplicado

# --- Nombres de Columnas (Lookup: LOOKUP_SUCURSAL_SHEET_NAME) ---
LKP_SUC_CLIENTE = 'Cliente'
LKP_SUC_SUCURSAL = 'Sucursal'
LKP_SUC_DEUDOR = 'Codigo Deudor'

# --- Nombres de Columnas (Lookup: LOOKUP_CSAP1_SHEET_NAME) ---
LKP_CSAP1_ITEM = 'Item'
LKP_CSAP1_CODIGO = 'CÓDIGO SAP'

# --- Nombres de Columnas (Lookup: LOOKUP_CSAP2_SHEET_NAME) ---
LKP_CSAP2_NOMBRE = 'Nombre de material'
LKP_CSAP2_CODIGO = 'Código SAP de Material2'
LKP_CSAP2_CORRIENTE = 'Corriente'

# --- Mapeo Directo de Columnas (Fuente -> Destino) ---
COLUMN_MAPPING_DIRECT = {
    SRC_COL_CLIENTE: TGT_COL_CLIENTE, SRC_COL_FECHA: TGT_COL_FECHA,
    SRC_COL_PLACA: TGT_COL_PLACA, SRC_COL_PESO: TGT_COL_PESO
}

# --- Columnas Adicionales a Procesar/Buscar en Hoja Destino ---
#    (Incluye todas las columnas que serán modificadas o buscadas)
EXTRA_TARGET_COLS_TO_PROCESS = [
    TGT_COL_SUCURSAL, TGT_COL_SAP, TGT_COL_DEUDOR_SUC,
    TGT_COL_CEDULA_CONDUCTOR, TGT_COL_CEDULA_AUXILIAR,
    TGT_COL_NOMBRE_ENTREGA, TGT_COL_CARGO_ENTREGA,
    TGT_COL_UNO # Se añade una vez, se manejará la duplicidad al buscar índices
]

# --- Configuraciones del Proceso ---
TARGET_START_ROW_NUM = 2
FUZZY_SAP_SIMILARITY_THRESHOLD = 90      # Umbral ajustado por el usuario
FUZZY_SUCURSAL_SIMILARITY_THRESHOLD = 90 # Umbral ajustado por el usuario
SKIPROWS_CSAP1 = 1
SKIPROWS_CSAP2 = 2
CONDUCTOR_GROUP_SIZE = 50 # Tamaño del grupo para cédulas (conductor y auxiliar)
AUXILIAR_GROUP_SIZE = 50 # Puede ser diferente si se desea

# --- Valores Fijos ---
FIXED_STRING_SIN_DESCRIPCION = "Sin descripción"
FIXED_NUMBER_UNO = 1

# --- Listas de Cédulas ---
LISTA_CEDULAS_CONDUCTOR = [
    '70328232', '71085884', '1128416491', '1000536279'
]
LISTA_CEDULAS_AUXILIAR = [ # <<< NUEVA LISTA >>>
    '1020394746', '1067863986', '98765808',
    '12021155', '7604671', '12021155' # Incluye duplicado como se proporcionó
]

# --- Estilos de Resaltado ---
HIGHLIGHT_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HIGHLIGHT_BLUE = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
HIGHLIGHT_NONE = PatternFill(fill_type=None)

# ==============================================================================
# --- Variables Globales ---
# ==============================================================================
selected_sap_code_from_popup: Optional[str] = None

# ==============================================================================
# --- Funciones de Limpieza de Texto ---
# (Sin cambios)
# ==============================================================================
def remove_accents(input_str: Any) -> str:
    if pd.isna(input_str): return ""
    try:
        text = str(input_str); nfkd_form = unicodedata.normalize('NFD', text)
        cleaned_text = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
        return cleaned_text.upper()
    except Exception as e:
        print(f"Advertencia(remove_accents): Error procesando '{input_str}': {e}")
        return str(input_str).upper()

def clean_text_for_comparison(text: Any) -> str:
    if pd.isna(text): return ""
    cleaned = remove_accents(str(text)); cleaned = cleaned.strip()
    return cleaned

def clean_client_name_part(client_name_part: str) -> str:
    if not client_name_part: return ""
    text = client_name_part.strip(); text = text.replace('.', '')
    text = remove_accents(text)
    suffixes = ['SAS', 'S A S', 'SA', 'S A', 'LTDA', 'LIMITADA', 'ESP', 'E S P']
    pattern = r'\s+\b(?:' + '|'.join(re.escape(s.replace('.', r'\.')) for s in suffixes) + r')\b$'
    text = re.sub(pattern, '', text, flags=re.IGNORECASE).strip()
    return text

def clean_sap_material_name(row: pd.Series, name_col: str, current_col: str) -> str:
    nombre = str(row[name_col]) if pd.notna(row[name_col]) else ""
    corriente = str(row[current_col]) if pd.notna(row[current_col]) else ""
    nombre = nombre.strip(); corriente = corriente.strip()
    cleaned_nombre = nombre
    if corriente:
        try:
            pattern = r'\s*' + re.escape(corriente) + r'\W*$'
            cleaned_nombre = re.sub(pattern, '', nombre, flags=re.IGNORECASE).strip()
        except Exception as e_regex: print(f"Advertencia(clean_sap): Regex falló n='{nombre}', c='{corriente}': {e_regex}")
    return remove_accents(cleaned_nombre)

def safe_str_conversion(value: Any) -> str:
    if pd.isna(value): return ""
    if isinstance(value, float) and not math.isnan(value) and value == int(value): return str(int(value))
    if isinstance(value, (int, float)) and not math.isnan(value): return str(value)
    return str(value).strip()

# ==============================================================================
# --- Funciones de Lectura y Preparación de Lookups ---
# (Sin cambios)
# ==============================================================================
def load_sucursal_lookup(excel_path: str) -> Tuple[Dict[str, str], List[Tuple[str, str, str]], List[str], Dict[str, Tuple[str, str]]]:
    sucursal_map_by_client = {}
    sucursal_deudor_data = []
    sucursal_names_clean = []
    deudor_map_self_match = {}
    try:
        df_suc = pd.read_excel(excel_path, sheet_name=LOOKUP_SUCURSAL_SHEET_NAME)
        print(f"Datos leídos de '{LOOKUP_SUCURSAL_SHEET_NAME}' ({len(df_suc)} filas).")
        df_suc.columns = df_suc.columns.str.strip()
        required_cols = [LKP_SUC_CLIENTE, LKP_SUC_SUCURSAL, LKP_SUC_DEUDOR]
        missing_cols = [col for col in required_cols if col not in df_suc.columns]
        if missing_cols: print(f"Advertencia: Faltan columnas en '{LOOKUP_SUCURSAL_SHEET_NAME}': {missing_cols}."); return {}, [], [], {}

        df_suc['cli_key_clean_part1'] = df_suc[LKP_SUC_CLIENTE].apply(lambda x: clean_client_name_part(str(x).split('-', maxsplit=1)[0]))
        df_suc['suc_key_clean'] = df_suc[LKP_SUC_SUCURSAL].apply(clean_text_for_comparison)
        df_suc['deudor_code_safe'] = df_suc[LKP_SUC_DEUDOR].apply(safe_str_conversion)
        df_suc['sucursal_original_safe'] = df_suc[LKP_SUC_SUCURSAL].apply(lambda x: str(x).strip())
        df_suc_valid = df_suc[(df_suc['deudor_code_safe'] != '') & (df_suc[LKP_SUC_SUCURSAL].notna()) & (df_suc['sucursal_original_safe'] != '')].copy()
        if df_suc_valid.empty: print("Advertencia: No se encontraron filas válidas en sucursales."); return {}, [], [], {}

        df_suc_valid['is_self_match'] = df_suc_valid.apply(lambda row: str(row[LKP_SUC_CLIENTE]).strip().upper() == str(row[LKP_SUC_SUCURSAL]).strip().upper(), axis=1)
        self_matching_rows = df_suc_valid[df_suc_valid['is_self_match']].copy()
        if not self_matching_rows.empty:
            self_matching_rows = self_matching_rows.drop_duplicates(subset=['cli_key_clean_part1'], keep='first')
            deudor_map_self_match = self_matching_rows.set_index('cli_key_clean_part1')[['deudor_code_safe', 'sucursal_original_safe']].apply(tuple, axis=1).to_dict()
            print(f" -> Mapa Prioritario (Cliente==Sucursal) creado con {len(deudor_map_self_match)} entradas.")
        else: print(" -> No se encontraron filas Cliente==Sucursal para mapa prioritario.")

        temp_map_fallback = df_suc_valid.dropna(subset=['cli_key_clean_part1', 'sucursal_original_safe'])
        sucursal_map_by_client = temp_map_fallback.drop_duplicates(subset=['cli_key_clean_part1'], keep='first').set_index('cli_key_clean_part1')['sucursal_original_safe'].to_dict()
        print(f" -> Mapa Fallback (Cliente P1->Sucursal) creado con {len(sucursal_map_by_client)} entradas.")

        fuzzy_match_rows = df_suc_valid[df_suc_valid['suc_key_clean'] != ''].copy()
        sucursal_deudor_data = list(fuzzy_match_rows[['suc_key_clean', 'deudor_code_safe', 'sucursal_original_safe']].itertuples(index=False, name=None))
        sucursal_names_clean = fuzzy_match_rows['suc_key_clean'].unique().tolist()
        print(f" -> Lista para Fuzzy Match (Sucursal P2) creada con {len(sucursal_deudor_data)} entradas.")
        print(f" -> Nombres de Sucursal únicos para Fuzzy Match: {len(sucursal_names_clean)}.")
    except FileNotFoundError: print(f"Error: No se encontró Excel en '{excel_path}'.")
    except ValueError as ve: print(f"Error: No se encontró hoja '{LOOKUP_SUCURSAL_SHEET_NAME}'. {ve}")
    except Exception as e: print(f"Error procesando '{LOOKUP_SUCURSAL_SHEET_NAME}': {e}"); traceback.print_exc()
    return sucursal_map_by_client, sucursal_deudor_data, sucursal_names_clean, deudor_map_self_match

def load_sap_lookups(excel_path: str) -> Tuple[Dict[str, str], List[str], Dict[str, str], List[str]]:
    codigosap_map1: Dict[str, str] = {}; csap1_keys: List[str] = []
    codigosap_map2: Dict[str, str] = {}; csap2_keys: List[str] = []
    try:
        df_csap1 = pd.read_excel(excel_path, sheet_name=LOOKUP_CSAP1_SHEET_NAME, skiprows=SKIPROWS_CSAP1)
        print(f"Datos leídos de '{LOOKUP_CSAP1_SHEET_NAME}' ({len(df_csap1)} filas).")
        df_csap1.columns = df_csap1.columns.str.strip()
        if LKP_CSAP1_ITEM in df_csap1.columns and LKP_CSAP1_CODIGO in df_csap1.columns:
            print(f"Preparando mapa SAP primario ('{LKP_CSAP1_ITEM}' -> '{LKP_CSAP1_CODIGO}')...")
            df_csap1['map_key'] = df_csap1[LKP_CSAP1_ITEM].apply(clean_text_for_comparison)
            df_csap1['sap_code_safe'] = df_csap1[LKP_CSAP1_CODIGO].apply(safe_str_conversion)
            temp_map1 = df_csap1[(df_csap1['map_key'] != '') & (df_csap1['sap_code_safe'] != '') & (df_csap1['sap_code_safe'].str.lower() != 'nan')].dropna(subset=['map_key', 'sap_code_safe'])
            codigosap_map1 = temp_map1.drop_duplicates(subset=['map_key'], keep='first').set_index('map_key')['sap_code_safe'].to_dict()
            csap1_keys = temp_map1['map_key'].unique().tolist()
            print(f" -> Mapa SAP primario ({len(codigosap_map1)}) y lista ({len(csap1_keys)}) creados.")
        else: print(f"Advertencia: Faltan '{LKP_CSAP1_ITEM}' o '{LKP_CSAP1_CODIGO}' en '{LOOKUP_CSAP1_SHEET_NAME}'.")
    except ValueError: print(f"Advertencia: No se encontró hoja '{LOOKUP_CSAP1_SHEET_NAME}'.")
    except Exception as e: print(f"Error procesando {LOOKUP_CSAP1_SHEET_NAME}: {e}")
    try:
        df_csap2 = pd.read_excel(excel_path, sheet_name=LOOKUP_CSAP2_SHEET_NAME, skiprows=SKIPROWS_CSAP2)
        print(f"Datos leídos de '{LOOKUP_CSAP2_SHEET_NAME}' ({len(df_csap2)} filas).")
        df_csap2.columns = df_csap2.columns.str.strip()
        required_cols = [LKP_CSAP2_NOMBRE, LKP_CSAP2_CODIGO, LKP_CSAP2_CORRIENTE]
        if all(col in df_csap2.columns for col in required_cols):
            print(f"Preparando mapa SAP secundario ('{LKP_CSAP2_NOMBRE}' limpio -> '{LKP_CSAP2_CODIGO}')...")
            df_csap2['sap_code_safe'] = df_csap2[LKP_CSAP2_CODIGO].apply(safe_str_conversion)
            df_csap2['map_key'] = df_csap2.apply(lambda row: clean_sap_material_name(row, LKP_CSAP2_NOMBRE, LKP_CSAP2_CORRIENTE), axis=1)
            temp_map2 = df_csap2[(df_csap2['map_key'] != '') & (df_csap2['sap_code_safe'] != '') & (df_csap2['sap_code_safe'].str.lower() != 'nan')].dropna(subset=['map_key', 'sap_code_safe'])
            codigosap_map2 = temp_map2.drop_duplicates(subset=['map_key'], keep='first').set_index('map_key')['sap_code_safe'].to_dict()
            csap2_keys = temp_map2['map_key'].unique().tolist()
            print(f" -> Mapa SAP secundario ({len(codigosap_map2)}) y lista ({len(csap2_keys)}) creados.")
        else: print(f"Advertencia: Faltan columnas {required_cols} en '{LOOKUP_CSAP2_SHEET_NAME}'.")
    except ValueError: print(f"Advertencia: No se encontró hoja '{LOOKUP_CSAP2_SHEET_NAME}'.")
    except Exception as e: print(f"Error procesando {LOOKUP_CSAP2_SHEET_NAME}: {e}")
    return codigosap_map1, csap1_keys, codigosap_map2, csap2_keys

# ==============================================================================
# --- Funciones Auxiliares del Proceso ---
# ==============================================================================
def prepare_sap_choices_for_popup(map1: Dict[str, str], map2: Dict[str, str]) -> List[str]:
    choices: Dict[str, str] = {}
    for item, code in map1.items():
        if item and code: display_text = f"{item} ({code})"; choices[display_text] = code
    for name, code in map2.items():
         if name and code: display_text = f"{name} ({code})";
         if display_text not in choices: choices[display_text] = code
    print(f"Lista de opciones SAP para popup generada con {len(choices)} entradas únicas.")
    return sorted(choices.keys())

def find_target_column_indices_with_duplicates(worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                               cols_to_find: List[str]) -> Dict[str, List[int]]:
    """
    Localiza los índices (número de columna) de las columnas requeridas en la hoja destino.
    Permite encontrar múltiples columnas con el mismo nombre (como '1').

    Args:
        worksheet: El objeto Worksheet de openpyxl para la hoja destino.
        cols_to_find: Lista de nombres de columnas destino a buscar.

    Returns:
        Un diccionario mapeando nombre de columna destino a una LISTA de sus índices.
        Ej: {'CLIENTE': [1], 'SUCURSAL': [2], '1': [5, 14]}
    """
    target_col_indices: Dict[str, List[int]] = {name: [] for name in cols_to_find}
    header_row = worksheet[1] # Asumir cabecera en la fila 1
    print(f"Localizando columnas destino en '{worksheet.title}' (Fila 1):")
    found_cols_report = []

    for cell in header_row:
        if cell.value:
            col_name = str(cell.value).strip()
            if col_name in target_col_indices:
                target_col_indices[col_name].append(cell.column)
                found_cols_report.append(f"'{col_name}' (Col {cell.column})")

    print(f"  - Encontradas: {', '.join(found_cols_report) if found_cols_report else 'Ninguna'}")

    # Reportar las que no se encontraron en absoluto
    missing_cols = [name for name, indices in target_col_indices.items() if not indices]
    if missing_cols:
        print(f"Advertencia: Columnas destino NO encontradas: {missing_cols}")

    return target_col_indices


def cleanup_rows_below_data(worksheet: openpyxl.worksheet.worksheet.Worksheet,
                            last_written_row: int,
                            processed_col_indices: List[int]):
    try:
        max_row_in_sheet = worksheet.max_row
        if not processed_col_indices: print("Advertencia(cleanup): No hay columnas para limpiar."); return
        if last_written_row < max_row_in_sheet:
            print(f"Limpiando datos/formato desde fila {last_written_row + 1} hasta {max_row_in_sheet}...")
            rows_cleared = 0
            for row_idx in range(last_written_row + 1, max_row_in_sheet + 1):
                row_had_data = False
                for col_idx in processed_col_indices:
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    has_fill = (hasattr(cell, 'fill') and cell.fill is not None and hasattr(cell.fill, 'patternType') and cell.fill.patternType is not None and cell.fill.patternType != 'none')
                    if not has_fill and hasattr(cell, 'fill') and cell.fill is not None: has_fill = (hasattr(cell.fill, 'fill_type') and cell.fill.fill_type is not None and cell.fill.fill_type != 'none')
                    if cell.value is not None or has_fill:
                        if not row_had_data: rows_cleared += 1; row_had_data = True
                        cell.value = None; cell.fill = HIGHLIGHT_NONE
            print(f"Limpieza de {rows_cleared} filas antiguas completada.")
        else: print("No se requiere limpieza de filas antiguas.")
    except Exception as e: print(f"Advertencia(cleanup): Error durante limpieza: {e}"); traceback.print_exc()

# ==============================================================================
# --- Interfaz Gráfica (Popup Selección Manual SAP) ---
# (Sin cambios)
# ==============================================================================
def show_manual_sap_selection_popup(parent_window: tk.Tk, residuo_name: str, sap_option_list: List[str]) -> Optional[str]:
    global selected_sap_code_from_popup; selected_sap_code_from_popup = "SKIP"
    try:
        popup = Toplevel(parent_window); popup.title("Selección Manual de Código SAP")
        wait_var = tk.IntVar(popup, value=0)
        window_width = 600; window_height = 500
        screen_width = popup.winfo_screenwidth(); screen_height = popup.winfo_screenheight()
        center_x = int(screen_width / 2 - window_width / 2); center_y = int(screen_height / 2 - window_height / 2)
        popup.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}'); popup.minsize(window_width, window_height)
        try: safe_residuo_name = residuo_name.encode(sys.getfilesystemencoding(), 'replace').decode(sys.getfilesystemencoding(), 'replace')
        except Exception: print(f"Advertencia: No se pudo codificar/decodificar residuo '{residuo_name}'."); safe_residuo_name = residuo_name
        main_label = Label(popup, text=f"Seleccione el código SAP para:\n'{safe_residuo_name}'", justify=tk.LEFT, wraplength=window_width - 20); main_label.pack(pady=(10, 5), padx=10, fill=tk.X)
        search_frame = Frame(popup); search_frame.pack(pady=5, padx=10, fill=tk.X)
        search_label = Label(search_frame, text="Buscar:"); search_label.pack(side=tk.LEFT, padx=(0, 5))
        search_var = StringVar(); search_entry = Entry(search_frame, textvariable=search_var, width=60); search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        list_frame = Frame(popup); list_frame.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)
        scrollbar = Scrollbar(list_frame, orient=tk.VERTICAL); scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        listbox = Listbox(list_frame, yscrollcommand=scrollbar.set, exportselection=False, selectmode=tk.SINGLE, activestyle='dotbox'); listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True); scrollbar.config(command=listbox.yview)
        def update_listbox_filter(search_term: str = ""):
            listbox.delete(0, tk.END); search_term_lower = search_term.lower(); items_added_count = 0
            for option_text in sap_option_list:
                if search_term_lower in option_text.lower():
                    try: safe_option = option_text.encode(sys.getfilesystemencoding(), 'replace').decode(sys.getfilesystemencoding(), 'replace'); listbox.insert(tk.END, safe_option); items_added_count += 1
                    except Exception as e: print(f"Advertencia(update_listbox): No se pudo añadir '{option_text}': {e}")
        def on_search_text_changed(*args): update_listbox_filter(search_var.get())
        search_var.trace_add("write", on_search_text_changed); update_listbox_filter()
        def handle_selection():
            global selected_sap_code_from_popup; selected_indices = listbox.curselection()
            if selected_indices:
                selected_text = listbox.get(selected_indices[0]); match = re.search(r'\(([^)]+)\)$', selected_text)
                if match: selected_sap_code_from_popup = match.group(1).strip()
                else: print(f"Advertencia: No se pudo extraer código SAP de '{selected_text}'."); selected_sap_code_from_popup = "SKIP"
                wait_var.set(1); popup.destroy()
            else: messagebox.showwarning("Sin Selección", "Seleccione un código o use 'Omitir'.", parent=popup)
        def handle_skip(): global selected_sap_code_from_popup; selected_sap_code_from_popup = "SKIP"; wait_var.set(1); popup.destroy()
        def handle_save_exit(): global selected_sap_code_from_popup; selected_sap_code_from_popup = "SAVE_EXIT"; wait_var.set(1); popup.destroy()
        def handle_window_close(): global selected_sap_code_from_popup; print("Advertencia: Popup SAP cerrado con 'X'."); selected_sap_code_from_popup = "SKIP"; wait_var.set(1); popup.destroy()
        popup.protocol("WM_DELETE_WINDOW", handle_window_close)
        button_frame = Frame(popup); button_frame.pack(pady=(10, 15))
        select_button = Button(button_frame, text="Seleccionar", command=handle_selection, width=15); select_button.pack(side=tk.LEFT, padx=5)
        skip_button = Button(button_frame, text="Omitir este Residuo", command=handle_skip, width=18); skip_button.pack(side=tk.LEFT, padx=5)
        save_exit_button = Button(button_frame, text="Guardar y Salir", command=handle_save_exit, width=15); save_exit_button.pack(side=tk.LEFT, padx=5)
        search_entry.focus_set(); popup.update_idletasks(); popup.lift()
        print("Esperando selección manual de SAP (wait_variable)...")
        popup.wait_variable(wait_var); print("... Selección SAP recibida o ventana cerrada.")
    except Exception as e_popup: print(f"ERROR FATAL creando/mostrando popup SAP: {e_popup}"); traceback.print_exc(); selected_sap_code_from_popup = None
    return selected_sap_code_from_popup

# ==============================================================================
# --- Función Principal de Ejecución ---
# ==============================================================================

def main():
    """Función principal que orquesta todo el proceso."""
    print("--- Iniciando Proceso de Procesamiento de Plantilla ---")
    root = tk.Tk(); root.withdraw()
    workbook: Optional[openpyxl.workbook.workbook.Workbook] = None
    input_excel_file: Optional[str] = None
    output_save_path: Optional[str] = None
    user_saved_mid_process: bool = False

    try:
        # --- 1. Selección Archivo ---
        print("\n[Paso 1/5] Selección del archivo Excel...")
        input_excel_file = filedialog.askopenfilename(parent=root, title="Selecciona el archivo Excel", filetypes=[("Excel", "*.xlsx"), ("*", "*")])
        if not input_excel_file: print("Operación cancelada."); return
        print(f"Archivo seleccionado: {input_excel_file}")

        # --- 2. Lectura y Preparación ---
        print(f"\n[Paso 2/5] Leyendo datos y preparando búsquedas...")
        try:
            df_source = pd.read_excel(input_excel_file, sheet_name=SOURCE_SHEET_NAME)
            print(f"Datos leídos de '{SOURCE_SHEET_NAME}' ({len(df_source)} filas).")
        except Exception as e: print(f"Error Crítico: No se pudo leer '{SOURCE_SHEET_NAME}'. {e}"); messagebox.showerror("Error Lectura", f"No se pudo leer '{SOURCE_SHEET_NAME}'.\n\n{e}"); return

        sucursal_map_fallback, suc_deudor_fuzzy_data, suc_names_fuzzy, deudor_map_priority = load_sucursal_lookup(input_excel_file)
        sap_map1, sap_keys1, sap_map2, sap_keys2 = load_sap_lookups(input_excel_file)
        sap_options_for_popup = prepare_sap_choices_for_popup(sap_map1, sap_map2)

        source_cols_required = list(COLUMN_MAPPING_DIRECT.keys()) + [SRC_COL_RESIDUO]
        missing_src_cols = [col for col in source_cols_required if col not in df_source.columns]
        if missing_src_cols: print(f"Error Crítico: Faltan columnas fuente: {missing_src_cols}"); messagebox.showerror("Error Columnas Fuente", f"Faltan en '{SOURCE_SHEET_NAME}':\n{', '.join(missing_src_cols)}"); return
        print("Verificación de columnas fuente: OK.")

        print("Pre-procesando claves de búsqueda...")
        df_source['lookup_key_client_part1'] = df_source[SRC_COL_CLIENTE].apply(lambda x: clean_client_name_part(str(x).split('-', maxsplit=1)[0]))
        df_source['lookup_key_residuo'] = df_source[SRC_COL_RESIDUO].apply(clean_text_for_comparison)
        def get_cleaned_part_after_hyphen(name): parts = str(name).split('-', 1); return clean_text_for_comparison(parts[1].strip()) if len(parts) > 1 else ""
        df_source['lookup_key_client_part2'] = df_source[SRC_COL_CLIENTE].apply(get_cleaned_part_after_hyphen)
        print("Pre-procesamiento de claves completado.")

        # --- Preparar listas de cédulas aleatorias ---
        driver_cedulas = LISTA_CEDULAS_CONDUCTOR[:] # Copiar
        if not driver_cedulas:
            print("Advertencia: Lista de cédulas de CONDUCTOR vacía.")
        else:
            random.shuffle(driver_cedulas)
            print(f"Lista de {len(driver_cedulas)} cédulas de CONDUCTOR mezclada (grupos de {CONDUCTOR_GROUP_SIZE}).")

        auxiliar_cedulas = LISTA_CEDULAS_AUXILIAR[:] # Copiar
        if not auxiliar_cedulas:
             print("Advertencia: Lista de cédulas de AUXILIAR vacía.")
        else:
            random.shuffle(auxiliar_cedulas)
            print(f"Lista de {len(auxiliar_cedulas)} cédulas de AUXILIAR mezclada (grupos de {AUXILIAR_GROUP_SIZE}).")


        # --- 3. Procesamiento Principal ---
        print(f"\n[Paso 3/5] Procesando filas y escribiendo en '{TARGET_SHEET_NAME}'...")
        try:
            workbook = openpyxl.load_workbook(input_excel_file)
            if TARGET_SHEET_NAME not in workbook.sheetnames: print(f"Error Crítico: No se encontró hoja destino '{TARGET_SHEET_NAME}'."); messagebox.showerror("Error Hoja Destino", f"No se encontró '{TARGET_SHEET_NAME}'."); return
            target_ws = workbook[TARGET_SHEET_NAME]
            print(f"Hoja destino '{TARGET_SHEET_NAME}' accesible.")
        except Exception as e: print(f"Error Crítico: No se pudo cargar/acceder a hoja destino. {e}"); messagebox.showerror("Error Workbook", f"No se pudo cargar/acceder a '{TARGET_SHEET_NAME}'.\n\n{e}"); return

        # <<< MODIFICADO: Usar la nueva función para obtener índices (maneja duplicados) >>>
        target_cols_to_find = list(COLUMN_MAPPING_DIRECT.values()) + EXTRA_TARGET_COLS_TO_PROCESS
        # Eliminar duplicados de la lista de búsqueda para evitar mensajes repetidos
        target_col_indices_map = find_target_column_indices_with_duplicates(target_ws, list(set(target_cols_to_find)))

        # Verificar columnas esenciales mapeadas
        missing_essential_tgt = [tgt for src, tgt in COLUMN_MAPPING_DIRECT.items() if not target_col_indices_map.get(tgt)]
        if missing_essential_tgt: print(f"Error Crítico: Faltan columnas destino mapeadas: {missing_essential_tgt}"); messagebox.showerror("Error Columnas Destino", f"Faltan en '{TARGET_SHEET_NAME}':\n{', '.join(missing_essential_tgt)}"); return

        # Obtener índices específicos (tomando el primero si hay duplicados, excepto para '1')
        tgt_idx_sucursal = target_col_indices_map.get(TGT_COL_SUCURSAL, [None])[0]
        tgt_idx_deudor_suc = target_col_indices_map.get(TGT_COL_DEUDOR_SUC, [None])[0]
        tgt_idx_sap = target_col_indices_map.get(TGT_COL_SAP, [None])[0]
        tgt_idx_cedula_conductor = target_col_indices_map.get(TGT_COL_CEDULA_CONDUCTOR, [None])[0]
        tgt_idx_cedula_auxiliar = target_col_indices_map.get(TGT_COL_CEDULA_AUXILIAR, [None])[0]
        tgt_idx_nombre_entrega = target_col_indices_map.get(TGT_COL_NOMBRE_ENTREGA, [None])[0]
        tgt_idx_cargo_entrega = target_col_indices_map.get(TGT_COL_CARGO_ENTREGA, [None])[0]
        # Obtener todos los índices para la columna '1'
        tgt_indices_uno = target_col_indices_map.get(TGT_COL_UNO, [])


        sap_items_for_manual_selection: List[Tuple[int, str]] = []
        processed_rows_count = 0
        last_written_excel_row = TARGET_START_ROW_NUM - 1

        # --- Bucle Principal ---
        for idx, source_row in df_source.iterrows():
            current_target_row = TARGET_START_ROW_NUM + idx
            last_written_excel_row = current_target_row; processed_rows_count += 1

            client_key_part1 = source_row['lookup_key_client_part1']
            client_key_part2 = source_row['lookup_key_client_part2']
            residuo_key = source_row['lookup_key_residuo']
            original_residuo_name = str(source_row[SRC_COL_RESIDUO]).strip()
            original_client_name = str(source_row[SRC_COL_CLIENTE]).strip()

            # --- a) Mapeo Directo ---
            for src_col, tgt_col in COLUMN_MAPPING_DIRECT.items():
                # Usar el primer índice encontrado para estas columnas
                tgt_idx = target_col_indices_map.get(tgt_col, [None])[0]
                if tgt_idx:
                    value_to_write = original_client_name if src_col == SRC_COL_CLIENTE else source_row[src_col]
                    if tgt_col == TGT_COL_FECHA and isinstance(value_to_write, pd.Timestamp): 
                        try: value_to_write = value_to_write.strftime('%Y-%m-%d') 
                        except ValueError: value_to_write = ""
                    elif tgt_col == TGT_COL_PESO: value_to_write = safe_str_conversion(value_to_write)
                    target_ws.cell(row=current_target_row, column=tgt_idx, value=value_to_write)

            # --- b) Búsqueda Sucursal / Código Deudor (PRIORIDAD INVERTIDA) ---
            final_sucursal_to_write = ""; final_deudor_code_to_write = ""
            match_found = False
            # Prioridad 1: Fuzzy Parte 2
            if not match_found and tgt_idx_deudor_suc and client_key_part2 and suc_names_fuzzy:
                match_info_part2 = process.extractOne(client_key_part2, suc_names_fuzzy, score_cutoff=FUZZY_SUCURSAL_SIMILARITY_THRESHOLD)
                if match_info_part2:
                    matched_suc_clean_name, score = match_info_part2
                    for suc_clean, debtor_code, suc_original in suc_deudor_fuzzy_data:
                        if suc_clean == matched_suc_clean_name:
                             final_deudor_code_to_write = debtor_code; final_sucursal_to_write = suc_original; match_found = True
                             print(f"  Fila {current_target_row}: Match Fuzzy (Parte 2: '{client_key_part2}') -> Suc: '{sucursal_original}', Deudor: '{debtor_code}' ({score}%)")
                             break
            # Prioridad 2: Cliente == Sucursal
            if not match_found and tgt_idx_deudor_suc and client_key_part1 in deudor_map_priority:
                 codigo_deudor, sucursal_original = deudor_map_priority[client_key_part1]
                 final_deudor_code_to_write = codigo_deudor; final_sucursal_to_write = sucursal_original; match_found = True
                 print(f"  Fila {current_target_row}: Match Prioritario (Cliente==Sucursal) -> Suc: '{sucursal_original}', Deudor: '{codigo_deudor}'")
            # Prioridad 3: Fallback
            if not match_found and tgt_idx_sucursal and sucursal_map_fallback:
                 sucursal_from_fallback = sucursal_map_fallback.get(client_key_part1, "")
                 if sucursal_from_fallback: final_sucursal_to_write = sucursal_from_fallback
            # Escribir resultados
            if tgt_idx_sucursal: target_ws.cell(row=current_target_row, column=tgt_idx_sucursal, value=final_sucursal_to_write)
            if tgt_idx_deudor_suc: target_ws.cell(row=current_target_row, column=tgt_idx_deudor_suc, value=final_deudor_code_to_write)

            # --- c) Búsqueda Código SAP ---
            found_sap_code = ""; sap_match_score = 0
            if tgt_idx_sap and (sap_keys1 or sap_keys2):
                sap_cell_to_write = target_ws.cell(row=current_target_row, column=tgt_idx_sap)
                if sap_keys1:
                    match1 = process.extractOne(residuo_key, sap_keys1, score_cutoff=FUZZY_SAP_SIMILARITY_THRESHOLD) # Corregido
                    if match1: key1, score1 = match1; found_sap_code = sap_map1.get(key1, ""); sap_match_score = score1
                if not found_sap_code and sap_keys2:
                    match2 = process.extractOne(residuo_key, sap_keys2, score_cutoff=FUZZY_SAP_SIMILARITY_THRESHOLD) # Corregido
                    if match2: key2, score2 = match2; found_sap_code = sap_map2.get(key2, ""); sap_match_score = score2
                if found_sap_code and str(found_sap_code).lower() != 'nan':
                    sap_cell_to_write.value = found_sap_code;
                    sap_cell_to_write.fill = HIGHLIGHT_YELLOW if FUZZY_SAP_SIMILARITY_THRESHOLD <= sap_match_score < 100 else HIGHLIGHT_NONE
                else:
                    sap_cell_to_write.value = None; sap_cell_to_write.fill = HIGHLIGHT_NONE;
                    sap_items_for_manual_selection.append((current_target_row, original_residuo_name))
            elif tgt_idx_sap: target_ws.cell(row=current_target_row, column=tgt_idx_sap, value=None).fill = HIGHLIGHT_NONE

            # --- d) Asignación Cédula Conductor ---
            if tgt_idx_cedula_conductor and driver_cedulas:
                group_index = idx // CONDUCTOR_GROUP_SIZE
                driver_index = group_index % len(driver_cedulas)
                assigned_cedula = driver_cedulas[driver_index]
                target_ws.cell(row=current_target_row, column=tgt_idx_cedula_conductor, value=assigned_cedula)
            elif tgt_idx_cedula_conductor: target_ws.cell(row=current_target_row, column=tgt_idx_cedula_conductor, value=None)

            # --- e) Asignación Cédula Auxiliar ---
            if tgt_idx_cedula_auxiliar and auxiliar_cedulas:
                group_index_aux = idx // AUXILIAR_GROUP_SIZE # Usar tamaño de grupo auxiliar
                aux_index = group_index_aux % len(auxiliar_cedulas)
                assigned_cedula_aux = auxiliar_cedulas[aux_index]
                target_ws.cell(row=current_target_row, column=tgt_idx_cedula_auxiliar, value=assigned_cedula_aux)
            elif tgt_idx_cedula_auxiliar: target_ws.cell(row=current_target_row, column=tgt_idx_cedula_auxiliar, value=None)

            # --- f) Asignación Valores Fijos ---
            if tgt_idx_nombre_entrega:
                target_ws.cell(row=current_target_row, column=tgt_idx_nombre_entrega, value=FIXED_STRING_SIN_DESCRIPCION)
            if tgt_idx_cargo_entrega:
                target_ws.cell(row=current_target_row, column=tgt_idx_cargo_entrega, value=FIXED_STRING_SIN_DESCRIPCION)

            # --- g) Asignación Columnas '1' ---
            # Escribir en todas las columnas encontradas que se llaman '1'
            if tgt_indices_uno:
                for col_1_idx in tgt_indices_uno:
                     target_ws.cell(row=current_target_row, column=col_1_idx, value=FIXED_NUMBER_UNO)

        # --- Fin del bucle for idx, source_row ---

        print(f"FASE 1 (Procesamiento Automático) completada. {processed_rows_count} filas procesadas.")
        if sap_items_for_manual_selection: print(f" -> {len(sap_items_for_manual_selection)} residuos requieren selección manual de SAP.")

        # --- 4. FASE 2: Selección Manual Interactiva (SAP) ---
        if sap_items_for_manual_selection and sap_options_for_popup and tgt_idx_sap:
            print(f"\n[Paso 4/5] Iniciando Selección Manual de SAP ({len(sap_items_for_manual_selection)} ítems)...")
            sap_selection_cache: Dict[str, str] = {}; selected_count, skipped_count, cache_hit_count = 0, 0, 0; abort_process = False
            for item_index, (target_row_num, residuo_name) in enumerate(sap_items_for_manual_selection):
                print(f"\nProc SAP Manual {item_index + 1}/{len(sap_items_for_manual_selection)} -> F:{target_row_num}, R:'{residuo_name}'")
                if residuo_name in sap_selection_cache: user_choice = sap_selection_cache[residuo_name]; print(f" -> Caché SAP: '{user_choice if user_choice != 'SKIP' else 'Omitido'}'"); cache_hit_count += 1
                else:
                    if not tk._default_root or not root.winfo_exists(): print("Adv: Recreando root Tk."); root = tk.Tk(); root.withdraw()
                    user_choice = show_manual_sap_selection_popup(root, residuo_name, sap_options_for_popup)
                    if user_choice is None: print("Error: Popup SAP falló."); abort_process = True; break
                    if user_choice == "SAVE_EXIT":
                        print(" -> 'Guardar y Salir'...");
                        if output_save_path is None:
                            base = os.path.splitext(os.path.basename(input_excel_file))[0]; sug = f"{base}_progreso.xlsx"; output_save_path = filedialog.asksaveasfilename(parent=root, title="Guardar Progreso", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile=sug)
                            if not output_save_path: print(" -> Guardado cancelado."); user_choice = "SKIP"
                            else: print(f" -> Path: {output_save_path}")
                        if output_save_path:
                            try:
                                print(f" -> Guardando: {output_save_path}..."); print(" -> Limpiando...");
                                # <<< Obtener TODOS los índices de columnas procesadas para limpieza >>>
                                all_processed_indices = []
                                for indices in target_col_indices_map.values():
                                    all_processed_indices.extend(indices)
                                cleanup_rows_below_data(target_ws, last_written_excel_row, list(set(all_processed_indices))) # Usar set para evitar duplicados
                                workbook.save(output_save_path); print(" -> Progreso guardado."); user_saved_mid_process = True; abort_process = True; break
                            except Exception as e: print(f" -> ERROR guardando: {e}"); messagebox.showerror("Error Guardar", f"No se pudo guardar.\n\n{e}", parent=root); user_choice = "SKIP"
                    if user_choice != "SAVE_EXIT": sap_selection_cache[residuo_name] = user_choice; (selected_count := selected_count + 1) if user_choice != "SKIP" else (skipped_count := skipped_count + 1)
                if user_choice not in ["SKIP", "SAVE_EXIT"] and tgt_idx_sap: cell = target_ws.cell(row=target_row_num, column=tgt_idx_sap); cell.value = user_choice; cell.fill = HIGHLIGHT_BLUE
                elif user_choice == "SKIP" and tgt_idx_sap: cell = target_ws.cell(row=target_row_num, column=tgt_idx_sap); cell.value = None; cell.fill = HIGHLIGHT_NONE
            if not abort_process: print(f"\nFASE 2 (SAP) Completada. Únicos: {len(sap_selection_cache)} (Sel:{selected_count}, Skip:{skipped_count}, Cache:{cache_hit_count})")
            else: print("\nFASE 2 (SAP) interrumpida.")
        elif sap_items_for_manual_selection: print("\n[Paso 4/5] Omitido: Selección Manual SAP no posible.")
        else: print("\n[Paso 4/5] Omitido: No hubo residuos para selección manual SAP.")


        # --- 5. Limpieza Final y Guardado ---
        if not user_saved_mid_process:
            print(f"\n[Paso 5/5] Limpiando y guardando...")
            # <<< Obtener TODOS los índices de columnas procesadas para limpieza final >>>
            all_processed_indices = []
            for indices in target_col_indices_map.values():
                all_processed_indices.extend(indices)
            cleanup_rows_below_data(target_ws, last_written_excel_row, list(set(all_processed_indices))) # Usar set para evitar duplicados

            base = os.path.splitext(os.path.basename(input_excel_file))[0]; sug = f"{base}_procesado.xlsx"; final_save_path = filedialog.asksaveasfilename(parent=root, title="Guardar archivo final", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile=sug)
            if not final_save_path: print("\nGuardado final cancelado.")
            else:
                try: workbook.save(final_save_path); print(f"\n--- ¡PROCESO COMPLETADO! ---"); print(f"Archivo guardado en: '{final_save_path}'"); messagebox.showinfo("Completado", f"Guardado en:\n{final_save_path}", parent=root)
                except Exception as e: print(f"\nError Crítico guardando final: {e}"); messagebox.showerror("Error Guardar Final", f"No se pudo guardar.\n\n{e}", parent=root)
        else: print(f"\n--- ¡PROCESO INTERRUMPIDO Y GUARDADO! ---"); print(f"Progreso guardado en: '{output_save_path}'"); messagebox.showinfo("Guardado", f"Progreso guardado en:\n{output_save_path}", parent=root)

    except Exception as main_error:
        print(f"\n--- ERROR INESPERADO ---"); print(f"Error: {main_error}"); traceback.print_exc()
        try: messagebox.showerror("Error Crítico", f"Error grave:\n\n{main_error}\n\nPrograma terminará.", parent=root)
        except Exception as e: print(f"Error adicional mostrando msg error: {e}")
    finally:
        if root:
            try: root.destroy()
            except Exception: pass
    print("\n--- Script finalizado. ---")

# ==============================================================================
# --- Punto de Entrada ---
# ==============================================================================
if __name__ == "__main__":
    main()